import os
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import sys
import json
import numpy as np
import cv2
from deskew import determine_skew

from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Tuple, Any

from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from PIL import Image


from collections import defaultdict
from openpyxl import Workbook, load_workbook
import copy
from copy import copy

import time
timings_ms = defaultdict(float)
timings_calls = defaultdict(int)
from contextlib import ContextDecorator
total_det_ms = 0.0

from pdf2image import convert_from_path, pdfinfo_from_path

import gc

from paddleocr import DocImgOrientationClassification, LayoutDetection, TextDetection, TextRecognition, TableRecognitionPipelineV2
import paddle


import bilstm_for_tables.inference as inference
from bilstm_for_tables.preprocessing import build_table_sample
from bilstm_for_tables.io_utils import load_document_annotation
from bilstm_for_tables.dataset import TableSequenceDataset

DEVICE = inference.DEVICE
NUM_BINS = inference.NUM_BINS
NUM_SLOTS = inference.NUM_SLOTS
VOCAB_PATH = inference.VOCAB_PATH

import bilstm_for_tables.config as config
ROW_CODES_PATH = config.ROW_CODES_PATH

BATCH_SIZE_TD  = 4   # LayoutDetection — по страницам
BATCH_SIZE_DET = 8   # TextDetection   — по страницам
BATCH_SIZE_REC = 8   # TextRecognition — по кропам
PDF_CONVERT_TREAD_COUNT = 8

MAX_PAGES = 99
RENDER_DPI = 200
MAX_LONG_SIDE = 1500

DEFAULT_NEW_BOX_CLASS_ID = 4
SCORE_THRESHOLD = 0.4

MIN_ABS_ANGLE = 0.99
MAX_ABS_ANGLE = 7.0

TARGET_CLASSES_FOR_TSR = {0, 1, 2, 3}

TABLE_PIPELINE = None
DOC_ORI_MODEL = None
OCR_DET_MODEL = None
OCR_REC_MODEL = None
LAYOUT_MODEL = None

TABLE_CLS_MODEL = None
TABLE_CLS_STOI = None
TABLE_CLS_ROW_CODES = None

@dataclass
class BBoxRel:
    x_min: float
    y_min: float
    x_max: float
    y_max: float

    def clipped(self) -> "BBoxRel":
        x1 = min(max(self.x_min, 0.0), 1.0)
        y1 = min(max(self.y_min, 0.0), 1.0)
        x2 = min(max(self.x_max, 0.0), 1.0)
        y2 = min(max(self.y_max, 0.0), 1.0)
        x1, x2 = sorted([x1, x2])
        y1, y2 = sorted([y1, y2])
        return BBoxRel(x1, y1, x2, y2)

    def to_list(self) -> List[float]:
        b = self.clipped()
        return [round(b.x_min, 6), round(b.y_min, 6), round(b.x_max, 6), round(b.y_max, 6)]

    @staticmethod
    def from_list(values: List[float]) -> "BBoxRel":
        return BBoxRel(*map(float, values)).clipped()


@dataclass
class TableAnnotation:
    table_id: int
    page_index: int
    class_id: int
    rotated_90: bool
    bbox: BBoxRel

@dataclass
class DocumentAnnotation:
    document_path: str
    num_pages: int
    tables: List[TableAnnotation] = field(default_factory=list)
    page_angles_deg: List[float] = field(default_factory=list)

@dataclass
class PageImage:
    page_index: int
    image_pil: Image.Image
    width_px: int
    height_px: int


def process_pdf_file(input_pdf_path: Path, output_xlsx_path: Path, working_dir: Path, log):

    log("⏳ Начало обработки файла")

    result = run_pipeline_incremental_until_ready(pdf_path=input_pdf_path, final_xlsx_path = output_xlsx_path, working_dir = working_dir)
    
    return result["xlsx_path"]



def get_model_dir(model_name: str):
    base_dir = Path(__file__).resolve().parent
    model_dir = base_dir / "models" / model_name

    if model_dir.exists() and model_dir.is_dir():
        return model_dir

    return None

def get_paddle_device():
    try:
        if paddle.device.is_compiled_with_cuda() and paddle.device.cuda.device_count() > 0:
            return "gpu:0"
    except Exception:
        pass
    return "cpu"

class time_it(ContextDecorator):
    def __init__(self, key: str):
        self.key = key
        self.start = None

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        elapsed_ms = (time.perf_counter() - self.start) * 1000
        timings_ms[self.key] += elapsed_ms
        timings_calls[self.key] += 1
        return False

def resize_keep_aspect(img: Image.Image, max_long_side: int) -> Image.Image:
    w, h = img.size
    long_side = max(w, h)
    if long_side <= max_long_side:
        return img

    scale = max_long_side / long_side
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))

    np_arr = np.array(img)
    resized_np = cv2.resize(
        np_arr, (new_w, new_h), interpolation=cv2.INTER_AREA  # ← лучше чем INTER_LINEAR для даунскейла
    )
    return Image.fromarray(resized_np)

def px_bbox_to_rel(bbox_px: List[float], width_px: int, height_px: int) -> BBoxRel:
    x1, y1, x2, y2 = bbox_px
    return BBoxRel(
        x_min=x1 / width_px,
        y_min=y1 / height_px,
        x_max=x2 / width_px,
        y_max=y2 / height_px,
    ).clipped()

def load_doc_orientation_model():
    global DOC_ORI_MODEL
    if DOC_ORI_MODEL is not None:
        return DOC_ORI_MODEL

    DOC_ORI_MODEL = DocImgOrientationClassification(model_name="PP-LCNet_x1_0_doc_ori", device=get_paddle_device(), model_dir = get_model_dir("PP-LCNet_x1_0_doc_ori"))
    DOC_ORI_MODEL.predict = time_it("ORI_predict")(DOC_ORI_MODEL.predict)
    return DOC_ORI_MODEL

def load_layout_model():
    global LAYOUT_MODEL
    if LAYOUT_MODEL is not None:
        return LAYOUT_MODEL

    LAYOUT_MODEL = LayoutDetection(model_name="PicoDet_layout_1x_table", device=get_paddle_device(), model_dir = get_model_dir("PicoDet_layout_1x_table"))
    LAYOUT_MODEL.predict = time_it("TD_predict")(LAYOUT_MODEL.predict)
    return LAYOUT_MODEL

def load_ocr_det_model():
    global OCR_DET_MODEL
    if OCR_DET_MODEL is not None:
        return OCR_DET_MODEL
    #OCR_DET_MODEL = TextDetection(model_name="PP-OCRv5_server_det")
    OCR_DET_MODEL = TextDetection(model_name="PP-OCRv5_mobile_det", device=get_paddle_device(), model_dir = get_model_dir("PP-OCRv5_mobile_det"))
    OCR_DET_MODEL.predict = time_it("DET_predict")(OCR_DET_MODEL.predict)
    return OCR_DET_MODEL


def load_ocr_rec_model():
    global OCR_REC_MODEL
    if OCR_REC_MODEL is not None:
        return OCR_REC_MODEL

    #OCR_REC_MODEL = TextRecognition(model_name="eslav_PP-OCRv5_mobile_rec", model_dir = r"C:\Users\GaV\.paddlex\official_models\eslav_PP-OCRv5_mobile_rec")
    OCR_REC_MODEL = TextRecognition(model_name="cyrillic_PP-OCRv3_mobile_rec", device=get_paddle_device(), model_dir = get_model_dir("cyrillic_PP-OCRv3_mobile_rec"))
    return OCR_REC_MODEL



def classify_page_orientations(images: List[Image.Image]) -> List[int]:
    """
    Возвращает список углов ориентации страниц в градусах (0, 90, 180, 270).
    Использует DocImgOrientationClassification из paddleocr.
    """
    if not images:
        return []

    model = load_doc_orientation_model()

    inputs: List[np.ndarray] = []
    for img in images:
        inputs.append(np.array(img.convert("RGB")))

    orientations: List[int] = []

    outputs = model.predict(input=inputs, batch_size=BATCH_SIZE_DET)

    class_id_to_angle = {
        0: 0,
        1: 90,
        2: 180,
        3: 270,
    }

    for out in outputs:
        angle = 0
        res = out.json["res"]
        label_names = res.get("label_names") or []

        if len(label_names) > 0:
            angle = int(str(label_names[0]).strip())
        elif "class_ids" in res:
            class_ids = res["class_ids"]
            class_id = int(class_ids[0])
            angle = class_id_to_angle.get(class_id, 0)

        if angle not in (0, 90, 180, 270):
            angle = 0

        orientations.append(angle)

    if len(orientations) != len(images):
        orientations = (
            orientations[: len(images)]
            + [0] * max(0, len(images) - len(orientations))
        )

    return orientations

def pdf_to_images(pdf_path: str, dpi: int = RENDER_DPI,
                max_long_side: int = MAX_LONG_SIDE,
                max_pages: int = MAX_PAGES,
                progress_callback=None,
                first_page: int = 1,
                last_page: Optional[int] = None) -> Tuple[List[PageImage], List[float]]:

    info = pdfinfo_from_path(pdf_path)
    page_count = info["Pages"]

    if last_page is None:
        last_page = page_count
    
    if first_page > page_count:
        return [], []
    
    last_page = min(last_page, page_count)
    selected_count = last_page - first_page + 1


    with time_it("Parallel PDF rendering"):
      raw_images = convert_from_path(
        pdf_path,
        dpi=dpi,
        fmt="RGB",
        thread_count=PDF_CONVERT_TREAD_COUNT,
        use_cropbox=True,
        first_page=first_page,
        last_page=last_page,
      )

    coarse_orientations = classify_page_orientations(raw_images)


    #  параллельная постобработка (выравнивание ориентации)
    def process_one(args):
      """Обрабатывает одну страницу. Возвращает (idx, PageImage, angle)."""
      idx, img, coarse_angle = args

      img = rotate_pil_by_orientation(img, coarse_angle)
      img, fine_angle = maybe_deskew_pil_image(img)

      img = resize_keep_aspect(img, max_long_side=max_long_side)

      global_page_index = first_page + idx

      page_image = PageImage(
            page_index = global_page_index,
            image_pil=img,
            width_px=img.width,
            height_px=img.height,
        )

      return idx, page_image, float(fine_angle)


    # Запускаем пул (результаты придут не в порядке — храним по idx)
    results: dict[int, tuple] = {}   # idx -> (PageImage, angle)
    completed = 0
    with time_it("Parallel render preprocessing"):
      with ThreadPoolExecutor(max_workers=PDF_CONVERT_TREAD_COUNT) as pool:
        futures = {
            pool.submit(process_one, (idx, img, coarse_orientations[idx])): idx
            for idx, img in enumerate(raw_images)
        }
        for future in as_completed(futures):
            idx, page_image, angle = future.result()
            results[idx] = (page_image, angle)
            completed += 1
            #if progress_callback is not None:
            #    progress_callback(completed, page_count, f"{completed}/{page_count}")


    # Восстанавливаем порядок страниц
    pages: List[PageImage] = []
    page_angles_deg: List[float] = []
    for idx in range(len(raw_images)):
        page_image, angle = results[idx]
        pages.append(page_image)
        page_angles_deg.append(angle)

    return pages, page_angles_deg


def rotate_pil(img: Image.Image, angle_deg: float) -> Image.Image:
    img_rgb = np.array(img.convert("RGB"))
    img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)

    h, w = img_bgr.shape[:2]
    center = (w / 2.0, h / 2.0)
    M = cv2.getRotationMatrix2D(center, angle_deg, 1.0)

    abs_cos = abs(M[0, 0])
    abs_sin = abs(M[0, 1])
    new_w = int(h * abs_sin + w * abs_cos)
    new_h = int(h * abs_cos + w * abs_sin)

    M[0, 2] += new_w / 2 - center[0]
    M[1, 2] += new_h / 2 - center[1]

    rotated_bgr = cv2.warpAffine(
        img_bgr,
        M,
        (new_w, new_h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255),
    )

    rotated_rgb = cv2.cvtColor(rotated_bgr, cv2.COLOR_BGR2RGB)
    return Image.fromarray(rotated_rgb)

def maybe_deskew_pil_image(img: Image.Image) -> tuple[Image.Image, float]:
    original_width, original_height = img.size

    scale = min(1.0, 700.0 / max(original_width, original_height))

    if scale < 1.0:
        img_small = img.resize(
            (int(original_width * scale), int(original_height * scale)),
            Image.LANCZOS
        )
    else:
        img_small = img

    np_small = np.array(img_small.convert("RGB"))
    img_bgr_small = cv2.cvtColor(np_small, cv2.COLOR_RGB2BGR)
    gray_small = cv2.cvtColor(img_bgr_small, cv2.COLOR_BGR2GRAY)

    angle = determine_skew(gray_small)

    if angle is None:
        return img, 0.0

    angle = float(angle)

    if abs(angle) < MIN_ABS_ANGLE or abs(angle) > MAX_ABS_ANGLE:
        return img, angle

    rotated_pil = rotate_pil(img, angle)

    return rotated_pil, angle

def rotate_pil_by_orientation(img: Image.Image, angle: int) -> Image.Image:
    if angle == 0:
        return img
    if angle == 90:
        return img.rotate(90, expand=True, fillcolor=(255, 255, 255))
    if angle == 180:
        return img
    if angle == 270:
        return img.rotate(270, expand=True, fillcolor=(255, 255, 255))
    return img

def run_tatr_on_document(page_images: List[PageImage], progress_callback=None) -> List[TableAnnotation]:
    """Детекция таблиц батчами по BATCH_SIZE_TD страниц."""
    all_tables: List[TableAnnotation] = []
    total = len(page_images)
    model = load_layout_model()

    next_table_id = 0

    for batch_start in range(0, total, BATCH_SIZE_TD):
        batch = page_images[batch_start : batch_start + BATCH_SIZE_TD]

        # Батч numpy-изображений
        np_images = [np.array(p.image_pil.convert("RGB")) for p in batch]
        outputs = model.predict(np_images, batch_size=len(np_images), layout_nms=True)

        for page_image, output in zip(batch, outputs):
            results = output['boxes']
            page_angle = 0 #int(orientation_by_page_index.get(page_image.page_index, 0))
            is_landscape = page_angle in (90, 270)

            for r in results:
                score = float(r.get("score", 1.0))
                if score < SCORE_THRESHOLD:
                    continue
                x1, y1, x2, y2 = r["coordinate"]
                bbox_rel = px_bbox_to_rel(
                    [float(x1), float(y1), float(x2), float(y2)],
                    page_image.width_px,
                    page_image.height_px,
                )
                all_tables.append(TableAnnotation(
                    table_id=next_table_id,
                    page_index=page_image.page_index,
                    class_id=DEFAULT_NEW_BOX_CLASS_ID,
                    rotated_90=is_landscape,
                    bbox=bbox_rel,
                ))
                next_table_id += 1

        if progress_callback is not None:
            done = min(batch_start + BATCH_SIZE_TD, total)
            progress_callback(done, total, f"Предразметка: {done}/{total}")

    return all_tables

def rel_bbox_to_px(bbox_rel, width_px, height_px):
    x1 = int(round(float(bbox_rel[0]) * width_px))
    y1 = int(round(float(bbox_rel[1]) * height_px))
    x2 = int(round(float(bbox_rel[2]) * width_px))
    y2 = int(round(float(bbox_rel[3]) * height_px))

    x1 = max(0, min(x1, width_px))
    y1 = max(0, min(y1, height_px))
    x2 = max(0, min(x2, width_px))
    y2 = max(0, min(y2, height_px))
    return [x1, y1, x2, y2]


def pad_images_to_same_shape(imgs):
    max_h = max(im.shape[0] for im in imgs)
    max_w = max(im.shape[1] for im in imgs)
    out = []
    for im in imgs:
        ph = max_h - im.shape[0]
        pw = max_w - im.shape[1]
        out.append(np.pad(im, ((0, ph), (0, pw), (0, 0)), constant_values=255))
    return out


def ensure_table_ids(base):
    tables = base.get("tables", [])
    for i, t in enumerate(tables, start=1):
        if "table_id" not in t or t["table_id"] is None:
            t["table_id"] = i
    return base


def build_empty_pages_branch(num_pages):
    return [{"page_index": i, "ocr_rows": []} for i in range(1, num_pages + 1)]


def run_ocr_det_on_table_crops(table_inputs):
    det_model = load_ocr_det_model()
    result = {}

    for batch_start in range(0, len(table_inputs), BATCH_SIZE_DET):
        batch = table_inputs[batch_start: batch_start + BATCH_SIZE_DET]
        batch_imgs = [x["crop_np"] for x in batch]
        batch_arr = pad_images_to_same_shape(batch_imgs)
        outputs = det_model.predict(batch_arr, batch_size=len(batch_arr))

        for item, output in zip(batch, outputs):
            raw_boxes = output["dt_polys"] if output is not None else []
            result[item["table_id"]] = list(raw_boxes)

    return result


def run_ocr_on_table_crop(page_image, table_id, table_bbox_rel, crop_pil, precomputed_boxes, run_rec=True):
    crop_w, crop_h = crop_pil.size
    raw_boxes = precomputed_boxes
    rows = []

    for row_idx, quad in enumerate(raw_boxes):
        quad = np.array(quad, dtype=np.float32)

        x1_crop = float(np.min(quad[:, 0]))
        y1_crop = float(np.min(quad[:, 1]))
        x2_crop = float(np.max(quad[:, 0]))
        y2_crop = float(np.max(quad[:, 1]))

        rows.append({
            "row_idx": row_idx,
            "table_id_prelabel": None,
            "table_id": int(table_id),
            "text": "",
            "text_conf": 0.0,
            "_bbox_crop_px": [x1_crop, y1_crop, x2_crop, y2_crop],
        })

    if run_rec and len(rows) > 0:
        rec_model = load_ocr_rec_model()
        crop_rgb = crop_pil.convert("RGB")

        valid_rows = []
        line_crops = []

        for row in rows:
            x1_px, y1_px, x2_px, y2_px = row["_bbox_crop_px"]
            x1_px = max(0, int(x1_px))
            y1_px = max(0, int(y1_px))
            x2_px = min(crop_w, int(x2_px))
            y2_px = min(crop_h, int(y2_px))
            if x2_px <= x1_px or y2_px <= y1_px:
                continue

            line_crop = np.array(crop_rgb.crop((x1_px, y1_px, x2_px, y2_px)))
            line_crops.append(line_crop)
            valid_rows.append(row)

        order = sorted(range(len(line_crops)), key=lambda i: line_crops[i].shape[1])
        line_crops = [line_crops[i] for i in order]
        valid_rows = [valid_rows[i] for i in order]

        for batch_start in range(0, len(line_crops), BATCH_SIZE_REC):
            batch_crops = line_crops[batch_start: batch_start + BATCH_SIZE_REC]
            batch_rows = valid_rows[batch_start: batch_start + BATCH_SIZE_REC]
            try:
                rec_outputs = rec_model.predict(batch_crops, batch_size=len(batch_crops))
                for row, rec_result in zip(batch_rows, rec_outputs):
                    if rec_result is not None:
                        row["text"] = str(rec_result.get("rec_text", "") or "")
                        row["text_conf"] = float(rec_result.get("rec_score", 0.0) or 0.0)
            except Exception as e:
                print(f"[ocr_rec][table_id={table_id}] batch_start={batch_start} error: {e}")

    table_x1_px, table_y1_px, table_x2_px, table_y2_px = rel_bbox_to_px(
        table_bbox_rel, page_image.width_px, page_image.height_px
    )

    final_rows = []
    for row in rows:
        x1_crop, y1_crop, x2_crop, y2_crop = row.pop("_bbox_crop_px")

        x1_page = table_x1_px + x1_crop
        y1_page = table_y1_px + y1_crop
        x2_page = table_x1_px + x2_crop
        y2_page = table_y1_px + y2_crop

        bbox_rel = px_bbox_to_rel(
            [x1_page, y1_page, x2_page, y2_page],
            page_image.width_px,
            page_image.height_px
        )

        final_rows.append({
            "row_idx": int(row["row_idx"]),
            "bbox": bbox_rel.to_list(),
            "table_id_prelabel": None,
            "table_id": int(row["table_id"]),
            "text": row["text"],
            "text_conf": float(row["text_conf"]),
        })

    return final_rows

def annotation_to_json(annotation: DocumentAnnotation) -> str:
    payload = {
        "document_path": annotation.document_path,
        "num_pages": int(annotation.num_pages),
        "tables": [
            {
                "table_id": int(t.table_id),
                "page_index": int(t.page_index),
                "class_id": int(t.class_id),
                "rotated_90": bool(t.rotated_90),
                "bbox": t.bbox.to_list(),
            }
            for t in annotation.tables
        ],
        "page_angles_deg": [float(a) for a in annotation.page_angles_deg],
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

def json_to_annotation(value: str) -> DocumentAnnotation:
    data = json.loads(value)

    tables = []
    for item in data.get("tables", []):
        tables.append(
            TableAnnotation(
                table_id=int(item["table_id"]),
                page_index=int(item["page_index"]),
                class_id=int(item["class_id"]),
                rotated_90=bool(item["rotated_90"]),
                bbox=BBoxRel.from_list(item["bbox"]),
            )
        )

    # page-wise углы поворота; для старых аннотаций поле может отсутствовать
    raw_angles = data.get("page_angles_deg", [])
    try:
        page_angles_deg = [float(a) for a in raw_angles]
    except Exception:
        page_angles_deg = []

    return DocumentAnnotation(
        document_path=str(data["document_path"]),
        num_pages=int(data["num_pages"]),
        tables=tables,
        page_angles_deg=page_angles_deg,
    )

def annotation_json_path(pdf_path: str) -> Path:
    """Путь к JSON-файлу аннотации рядом с PDF."""
    return Path(pdf_path).with_suffix(".annotation.json")

def save_annotation_to_file(pdf_path: str, annotation_json: str) -> None:
    """Атомарно сохраняет JSON в файл рядом с PDF."""
    target = annotation_json_path(pdf_path)
    tmp = target.with_suffix(".tmp.json")
    tmp.write_text(annotation_json, encoding="utf-8")
    tmp.replace(target)

def build_annotation_with_table_ocr(pdf_path, run_rec=True, first_page=1, last_page=None,):

    pages, page_angles_deg = pdf_to_images(
        pdf_path,
        dpi=RENDER_DPI,
        max_long_side=MAX_LONG_SIDE,
        max_pages=MAX_PAGES,
        progress_callback=None,
        first_page=first_page,
        last_page=last_page,
    )

    if not pages:
        info = pdfinfo_from_path(pdf_path)
        total_pages = int(info["Pages"])

        base = {
            "document_path": pdf_path,
            "num_pages": total_pages,
            "tables": [],
            "page_angles_deg": [0.0] * total_pages,
            "logical_tables": [],
            "pages": build_empty_pages_branch(total_pages),
        }
        return base, pages

    info = pdfinfo_from_path(pdf_path)
    total_pages = int(info["Pages"])

    tables = run_tatr_on_document(pages, progress_callback=None)

    annotation = DocumentAnnotation(
        document_path=pdf_path,
        num_pages=total_pages,
        tables=tables,
        page_angles_deg=[0.0] * total_pages,
    )

    for page, angle in zip(pages, page_angles_deg):
        page_idx_0 = int(page.page_index) - 1
        if 0 <= page_idx_0 < total_pages:
            annotation.page_angles_deg[page_idx_0] = float(angle)

    base = json.loads(annotation_to_json(annotation))
    base = ensure_table_ids(base)

    pages_by_index = {int(p.page_index): p for p in pages}

    pages_out = build_empty_pages_branch(total_pages)
    pages_out_by_index = {int(x["page_index"]): x for x in pages_out}

    table_inputs = []

    for t in base["tables"]:
        page_index = int(t["page_index"])
        table_id = int(t["table_id"])

        page_image = pages_by_index.get(page_index)
        if page_image is None:
            continue

        real_bbox_rel = list(t["bbox"])
        ocr_bbox_rel = [float(real_bbox_rel[0]), max(0.0, float(real_bbox_rel[1]) - 0.10), float(real_bbox_rel[2]), float(real_bbox_rel[3])]

        x1, y1, x2, y2 = rel_bbox_to_px(ocr_bbox_rel, page_image.width_px, page_image.height_px)

        if x2 <= x1 or y2 <= y1:
            continue

        crop_pil = page_image.image_pil.convert("RGB").crop((x1, y1, x2, y2))
        crop_np = np.array(crop_pil)

        if crop_np.size == 0:
            continue

        table_inputs.append(
            {
                "table_id": table_id,
                "page_index": page_index,
                "bbox_rel": real_bbox_rel,
                "ocr_bbox_rel": ocr_bbox_rel,
                "crop_pil": crop_pil,
                "crop_np": crop_np,
            }
        )

    det_boxes_by_table_id = (
        run_ocr_det_on_table_crops(table_inputs)
        if len(table_inputs) > 0
        else {}
    )

    for item in table_inputs:
        table_id = item["table_id"]
        page_index = item["page_index"]
        page_image = pages_by_index[page_index]
        raw_boxes = det_boxes_by_table_id.get(table_id, [])

        table_rows = run_ocr_on_table_crop(
            page_image=page_image,
            table_id=table_id,
            table_bbox_rel=item["ocr_bbox_rel"],
            crop_pil=item["crop_pil"],
            precomputed_boxes=raw_boxes,
            run_rec=run_rec,
        )

        pages_out_by_index[page_index]["ocr_rows"].extend(table_rows)

    for page_index in pages_out_by_index:
        rows = pages_out_by_index[page_index]["ocr_rows"]
        rows.sort(key=lambda r: (r["bbox"][1], r["bbox"][0]))
        for new_idx, row in enumerate(rows):
            row["row_idx"] = new_idx

    base["pages"] = pages_out
    return base, pages


def annotation_json_path(pdf_path: str) -> Path:
    return Path(pdf_path).with_suffix(".annotation.json")


def load_table_classifier_stoi_once():
    global TABLE_CLS_STOI
    if TABLE_CLS_STOI is None:
        _, TABLE_CLS_STOI = inference.load_vocab(VOCAB_PATH)
    return TABLE_CLS_STOI


def load_table_classifier_row_codes_once():
    global TABLE_CLS_ROW_CODES
    if TABLE_CLS_ROW_CODES is None:
        row_codes_path = ROW_CODES_PATH
        with open(row_codes_path, "r", encoding="utf-8") as f:
            TABLE_CLS_ROW_CODES = [line.rstrip("\n") for line in f]
    return TABLE_CLS_ROW_CODES


def load_table_classifier_model_once(samples_for_probe, stoi):
    global TABLE_CLS_MODEL

    if TABLE_CLS_MODEL is None:
        if not samples_for_probe:
            raise ValueError("samples_for_probe must not be empty when initializing classification model")


        probe_ds = TableSequenceDataset(
            samples=samples_for_probe,
            stoi=stoi,
            max_text_len=inference.MAX_TEXT_LEN,
        )
        TABLE_CLS_MODEL = inference.build_model(
            stoi=stoi,
            num_numeric_features=probe_ds.num_numeric_features,
        )
        TABLE_CLS_MODEL.eval()

    return TABLE_CLS_MODEL


def build_samples_for_document(doc, row_codes):
    samples = []
    for table in doc.tables:
        sample = build_table_sample(
            doc=doc,
            table=table,
            row_codes=row_codes,
            num_bins=NUM_BINS,
            num_slots=NUM_SLOTS,
        )
        if sample is not None:
            samples.append(sample)
    return samples

@time_it("CLS_predict")
def predict_table_classes_from_json_path(json_path):
    doc = load_document_annotation(str(json_path))

    row_codes = load_table_classifier_row_codes_once()
    samples = build_samples_for_document(doc, row_codes)

    if not samples:
        return []

    stoi = load_table_classifier_stoi_once()
    model = load_table_classifier_model_once(samples_for_probe=samples, stoi=stoi)

    cls_results = inference.predict_document_samples(
        model=model,
        samples=samples,
        stoi=stoi,
    )

    for item in cls_results:
        item["page_index_zero_based"] = int(item["page_index"])
        item["page_index"] = int(item["page_index"]) + 1

    return cls_results


def apply_table_classification_to_annotation_dict(annotation_dict, cls_results):
    pred_by_table_id = {}
    for item in cls_results:
        pred_by_table_id[int(item["table_id"])] = {
            "pred_class": int(item["pred_class"]),
            "pred_confidence": float(item["pred_confidence"]),
        }

    for t in annotation_dict.get("tables", []):
        table_id = int(t["table_id"])
        if table_id in pred_by_table_id:
            t["class_id"] = int(pred_by_table_id[table_id]["pred_class"])
            # t["class_conf"] = float(pred_by_table_id[table_id]["pred_confidence"])

    return annotation_dict


def run_table_classification_step(pdf_path, annotation_dict=None, save_json=True):
    json_path = annotation_json_path(pdf_path)

    if annotation_dict is not None:
        json_path.write_text(
            json.dumps(annotation_dict, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8"
        )

    if not json_path.exists():
        raise FileNotFoundError(f"Не найден annotation json: {json_path}")

    cls_results = predict_table_classes_from_json_path(json_path)

    annotation_dict_loaded = json.loads(json_path.read_text(encoding="utf-8"))
    annotation_dict_loaded = apply_table_classification_to_annotation_dict(annotation_dict_loaded, cls_results)

    if save_json:
        json_path.write_text(
            json.dumps(annotation_dict_loaded, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8"
        )

    return annotation_dict_loaded, cls_results

def crop_table_image(page_image, bbox_rel):
    x1, y1, x2, y2 = rel_bbox_to_px(bbox_rel, page_image.width_px, page_image.height_px)
    if x2 <= x1 or y2 <= y1:
        return None
    crop = page_image.image_pil.convert("RGB").crop((x1, y1, x2, y2))
    if crop.size[0] <= 0 or crop.size[1] <= 0:
        return None
    return crop


def resize_to_width_keep_aspect(img, target_width):
    if img is None:
        return None
    w, h = img.size
    if w <= 0 or h <= 0 or target_width <= 0:
        return None
    if w == target_width:
        return img
    new_h = max(1, int(round(h * target_width / w)))
    return img.resize((target_width, new_h), Image.Resampling.LANCZOS)


def stack_images_vertically(images, bg_color=(255, 255, 255)):
    images = [img for img in images if img is not None]
    if not images:
        return None

    target_width = max(img.size[0] for img in images)
    resized = [resize_to_width_keep_aspect(img, target_width) for img in images]
    resized = [img for img in resized if img is not None]

    total_h = sum(img.size[1] for img in resized)
    canvas = Image.new("RGB", (target_width, total_h), color=bg_color)

    y = 0
    for img in resized:
        canvas.paste(img, (0, y))
        y += img.size[1]

    return canvas


def collect_target_tables(annotation_dict):
    tables = []
    for t in annotation_dict.get("tables", []):
        class_id = int(t.get("class_id", 4))
        if class_id in TARGET_CLASSES_FOR_TSR:
            tables.append({
                "table_id": int(t["table_id"]),
                "page_index": int(t["page_index"]),   # 1-based
                "class_id": class_id,
                "bbox": list(t["bbox"]),
                "rotated_90": bool(t.get("rotated_90", False)),
            })
    return tables


def build_conservative_table_groups(annotation_dict):
    tables = collect_target_tables(annotation_dict)

    by_class = {}
    for t in tables:
        by_class.setdefault(t["class_id"], []).append(t)

    final_groups = []

    for class_id, class_tables in by_class.items():
        page_to_tables = {}
        for t in class_tables:
            page_to_tables.setdefault(t["page_index"], []).append(t)

        candidate_pages = sorted(
            p for p, items in page_to_tables.items()
            if len(items) == 1
        )

        runs = []
        current_run = []

        for p in candidate_pages:
            if not current_run:
                current_run = [p]
            elif p == current_run[-1] + 1:
                current_run.append(p)
            else:
                runs.append(current_run)
                current_run = [p]

        if current_run:
            runs.append(current_run)

        pages_in_runs = set()
        for run in runs:
            if len(run) >= 2:
                group_tables = [page_to_tables[p][0] for p in run]
                final_groups.append({
                    "group_type": "merged_sequence",
                    "class_id": class_id,
                    "page_indices": run,
                    "tables": group_tables,
                })
                pages_in_runs.update(run)

        for p, items in sorted(page_to_tables.items()):
            if p in pages_in_runs:
                continue
            if len(items) == 1:
                final_groups.append({
                    "group_type": "single_table",
                    "class_id": class_id,
                    "page_indices": [p],
                    "tables": [items[0]],
                })
            else:
                for t in items:
                    final_groups.append({
                        "group_type": "single_table_ambiguous_page",
                        "class_id": class_id,
                        "page_indices": [p],
                        "tables": [t],
                    })

    final_groups.sort(key=lambda g: (g["page_indices"][0], g["class_id"], g["tables"][0]["table_id"]))
    return final_groups


def build_tsr_candidates(annotation_dict, pages):
    pages_by_index = {int(p.page_index): p for p in pages}
    groups = build_conservative_table_groups(annotation_dict)

    candidates = []

    for group_idx, group in enumerate(groups):
        class_id = int(group["class_id"])
        group_type = group["group_type"]
        tables = group["tables"]

        table_crops = []
        valid_tables = []

        for t in tables:
            page_index = int(t["page_index"])
            page_image = pages_by_index.get(page_index)
            if page_image is None:
                continue

            crop = crop_table_image(page_image, t["bbox"])
            if crop is None:
                continue

            table_crops.append(crop)
            valid_tables.append(t)

        if not table_crops:
            continue

        if len(table_crops) == 1:
            final_image = table_crops[0]
        else:
            final_image = stack_images_vertically(table_crops)

        candidates.append({
            "candidate_id": group_idx,
            "class_id": class_id,
            "group_type": group_type,
            "page_indices": [int(t["page_index"]) for t in valid_tables],
            "table_ids": [int(t["table_id"]) for t in valid_tables],
            "source_tables": valid_tables,
            "image_pil": final_image,
        })

    return candidates

def load_table_pipeline():
    global TABLE_PIPELINE
    if TABLE_PIPELINE is not None:
        return TABLE_PIPELINE

    TABLE_PIPELINE = TableRecognitionPipelineV2(
        device= get_paddle_device(),
        table_classification_model_name="PP-LCNet_x1_0_table_cls",
        table_classification_model_dir = get_model_dir("PP-LCNet_x1_0_table_cls"), 
        text_detection_model_name="PP-OCRv5_mobile_det",
        text_detection_model_dir = get_model_dir("PP-OCRv5_mobile_det"),
        text_recognition_model_name= "cyrillic_PP-OCRv5_mobile_rec", #"eslav_PP-OCRv5_mobile_rec",
        text_recognition_model_dir = get_model_dir("cyrillic_PP-OCRv5_mobile_rec"),
        text_recognition_batch_size=16,
        wired_table_cells_detection_model_name="RT-DETR-L_wired_table_cell_det",
        wired_table_cells_detection_model_dir = get_model_dir("RT-DETR-L_wired_table_cell_det"),
        wireless_table_cells_detection_model_name="RT-DETR-L_wireless_table_cell_det",
        wireless_table_cells_detection_model_dir=get_model_dir("RT-DETR-L_wireless_table_cell_det"),
        #table_algorithm="TableMaster",
        wired_table_structure_recognition_model_name="SLANeXt_wired",
        wired_table_structure_recognition_model_dir=get_model_dir("SLANeXt_wired"),
        wireless_table_structure_recognition_model_name="SLANet_plus",
        wireless_table_structure_recognition_model_dir = get_model_dir("SLANet_plus"),
        doc_orientation_classify_model_name = "PP-LCNet_x1_0_doc_ori",
        doc_orientation_classify_model_dir= get_model_dir("PP-LCNet_x1_0_doc_ori"), 
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_layout_detection=False,
        use_ocr_model=True,
    )
    return TABLE_PIPELINE

def run_table_pipeline_on_candidates(tsr_candidates):
    pipeline = load_table_pipeline()

    imgs = [np.array(c["image_pil"].convert("RGB")) for c in tsr_candidates]
    outputs = pipeline.predict(imgs)

    return outputs

def export_table_pipeline_results_to_excel(
    tsr_candidates,
    pipeline_results,
    pdf_path,
    final_xlsx_path,
    working_dir
):
    class_to_sheet = {
        0: "ОФП",
        1: "ОФР",
        2: "ОДДС",
        3: "ОИК",
    }


    saved_files = defaultdict(list)

    for cand, res in zip(tsr_candidates, pipeline_results):
        class_id = int(cand["class_id"])
        if class_id not in class_to_sheet:
            continue

        class_dir = Path(working_dir) / f"class_{class_id}"
        class_dir.mkdir(parents=True, exist_ok=True)

        candidate_id = cand["candidate_id"]
        out_name = f"candidate_{candidate_id}.xlsx"
        out_path = class_dir / out_name

        try:
            res.save_to_xlsx(str(out_path))
        except Exception as e:
            print(f"[WARN] failed to save candidate_id={candidate_id} to xlsx: {e}")
            continue

        saved_files[class_id].append((cand, out_path))

    total_saved = sum(len(v) for v in saved_files.values())
    if total_saved == 0:
        print("Нет таблиц для записи в Excel.")
        return None

    #final_xlsx_path = output_dir / f"{Path(pdf_path).stem}_tables_pipeline_v2.xlsx"

    out_wb = Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    out_ws_map = {}
    next_row_map = {}

    for class_id in sorted(saved_files.keys()):
        sheet_name = class_to_sheet[class_id]
        ws = out_wb.create_sheet(title=sheet_name)
        out_ws_map[class_id] = ws
        next_row_map[class_id] = 1

    def copy_sheet_content(src_ws, dst_ws, start_row):
        row_offset = start_row - 1

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = dst_ws.cell(
                    row=cell.row + row_offset,
                    column=cell.column,
                    value=cell.value,
                )

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)

        for col_letter, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_letter].width = dim.width

        for row_idx, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[row_idx + row_offset].height = dim.height

        merged_ranges = list(src_ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row + row_offset
            max_col = merged_range.max_col
            max_row = merged_range.max_row + row_offset
            dst_ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

        return start_row + src_ws.max_row

    for class_id in sorted(saved_files.keys()):
        dst_ws = out_ws_map[class_id]

        for cand, xlsx_path in saved_files[class_id]:
            try:
                src_wb = load_workbook(xlsx_path)
                src_ws = src_wb[src_wb.sheetnames[0]]
            except Exception as e:
                print(f"[WARN] failed to read intermediate xlsx {xlsx_path}: {e}")
                continue


            cur_row = next_row_map[class_id]

            cur_row = copy_sheet_content(src_ws, dst_ws, cur_row)
            cur_row += 2

            next_row_map[class_id] = cur_row

    out_wb.save(final_xlsx_path)

    print(f"Сохранено Excel-файлов: {total_saved}")
    print(f"Итоговый Excel сохранён: {final_xlsx_path}")
    return final_xlsx_path


def merge_annotation_chunks(
    accumulated_annotation: Optional[dict],
    accumulated_pages: Optional[dict[int, PageImage]],
    annotation_chunk: dict,
    pages_chunk: list[PageImage],
    next_table_id: int,
) -> tuple[dict, dict[int, PageImage], int]:
    """
    Merge one processed PDF chunk into accumulated document state.

    Parameters
    ----------
    accumulated_annotation
        Previously accumulated annotationdict for the whole document.
        Can be None on the first chunk.
    accumulated_pages
        Previously cached pages by page_index. Can be None on the first chunk.
    annotation_chunk
        Annotation dict returned by chunk processing for current page window.
    pages_chunk
        Rendered PageImage objects for current page window.
    next_table_id
        Global next free table id for the whole document.

    Returns
    -------
    tuple[dict, dict[int, PageImage], int]
        (merged_annotation, merged_pages_by_index, updated_next_table_id)
    """
    if accumulated_annotation is None:
        accumulated_annotation = {
            "document_path": str(annotation_chunk.get("document_path", "")),
            "num_pages": int(annotation_chunk.get("num_pages", 0)),
            "tables": [],
            "page_angles_deg": list(annotation_chunk.get("page_angles_deg", [])),
            "pages": [],
            "status": {},
            "logical_tables": [],
        }
    else:
        pass
        #accumulated_annotation = json.loads(
        #    json.dumps(accumulated_annotation, ensure_ascii=False)
        #)

    if accumulated_pages is None:
        accumulated_pages = {}

    accumulated_annotation.setdefault("tables", [])
    accumulated_annotation.setdefault("pages", [])
    accumulated_annotation.setdefault("status", {})
    accumulated_annotation.setdefault("logical_tables", [])
    accumulated_annotation.setdefault("page_angles_deg", [])

    chunk_tables = list(annotation_chunk.get("tables", []))
    chunk_pages_meta = list(annotation_chunk.get("pages", []))
    chunk_page_angles = list(annotation_chunk.get("page_angles_deg", []))
    chunk_num_pages = int(annotation_chunk.get("num_pages", 0))

    merged_num_pages = max(
        int(accumulated_annotation.get("num_pages", 0)),
        chunk_num_pages,
    )
    accumulated_annotation["num_pages"] = merged_num_pages

    existing_angles = list(accumulated_annotation.get("page_angles_deg", []))
    if len(existing_angles) < merged_num_pages:
        existing_angles.extend([0.0] * (merged_num_pages - len(existing_angles)))

    for idx, angle in enumerate(chunk_page_angles, start=1):
        if 1 <= idx <= merged_num_pages:
            existing_angles[idx - 1] = float(angle)

    accumulated_annotation["page_angles_deg"] = existing_angles

    pages_meta_by_index: dict[int, dict] = {}
    for page_meta in accumulated_annotation.get("pages", []):
        page_index = int(page_meta["page_index"])
        pages_meta_by_index[page_index] = page_meta

    table_id_mapping: dict[int, int] = {}

    for table in chunk_tables:
        old_table_id = int(table["table_id"])
        new_table_id = int(next_table_id)
        next_table_id += 1
        table_id_mapping[old_table_id] = new_table_id

        table_copy = dict(table)
        table_copy["table_id"] = new_table_id
        accumulated_annotation["tables"].append(table_copy)

    for page_meta in chunk_pages_meta:
        page_index = int(page_meta["page_index"])
        rows = list(page_meta.get("ocr_rows", []))

        remapped_rows = []
        for row in rows:
            row_copy = dict(row)
            old_table_id = row_copy.get("table_id")
            old_table_id_prelabel = row_copy.get("table_id_prelabel")

            if old_table_id is not None:
                row_copy["table_id"] = table_id_mapping.get(int(old_table_id), int(old_table_id))

            if old_table_id_prelabel is not None:
                row_copy["table_id_prelabel"] = table_id_mapping.get(
                    int(old_table_id_prelabel),
                    int(old_table_id_prelabel),
                )

            remapped_rows.append(row_copy)

        if page_index not in pages_meta_by_index:
            pages_meta_by_index[page_index] = {
                "page_index": page_index,
                "ocr_rows": remapped_rows,
            }
        else:
            pages_meta_by_index[page_index].setdefault("ocr_rows", [])
            pages_meta_by_index[page_index]["ocr_rows"].extend(remapped_rows)

    merged_pages_meta = []
    for page_index in sorted(pages_meta_by_index.keys()):
        page_meta = pages_meta_by_index[page_index]
        page_meta.setdefault("ocr_rows", [])
        page_meta["ocr_rows"].sort(
            key=lambda row: (
                float(row["bbox"][1]),
                float(row["bbox"][0]),
            )
        )
        for new_row_idx, row in enumerate(page_meta["ocr_rows"]):
            row["row_idx"] = int(new_row_idx)
        merged_pages_meta.append(page_meta)

    accumulated_annotation["pages"] = merged_pages_meta

    for page in pages_chunk:
        accumulated_pages[int(page.page_index)] = page

    return accumulated_annotation, accumulated_pages, next_table_id


def check_tsr_ready_conditions(
    annotation_dict: dict,
    last_processed_page: int,
    target_classes: tuple[int, ...] = (0, 1, 2, 3),
) -> dict:
    """
    Check whether it is safe to stop early-stage scanning and start TSR.

    Conditions:
    1) all target classes are found
    2) no target-class table is located on the last processed page
    """
    tables = list(annotation_dict.get("tables", []))
    target_classes_set = set(int(x) for x in target_classes)

    target_tables = []
    found_classes = set()
    target_tables_on_last_page = []

    for table in tables:
        class_id = int(table.get("class_id", -1))
        if class_id not in target_classes_set:
            continue

        page_index = int(table.get("page_index", -1))
        table_id = int(table.get("table_id", -1))

        target_tables.append(table)
        found_classes.add(class_id)

        if page_index == int(last_processed_page):
            target_tables_on_last_page.append(
                {
                    "table_id": table_id,
                    "page_index": page_index,
                    "class_id": class_id,
                }
            )

    missing_classes = sorted(target_classes_set - found_classes)

    all_target_classes_found = len(missing_classes) == 0
    no_target_tables_on_last_page = len(target_tables_on_last_page) == 0
    ready_for_tsr = all_target_classes_found and no_target_tables_on_last_page

    return {
        "ready_for_tsr": bool(ready_for_tsr),
        "all_target_classes_found": bool(all_target_classes_found),
        "no_target_tables_on_last_page": bool(no_target_tables_on_last_page),
        "found_classes": sorted(found_classes),
        "missing_classes": missing_classes,
        "last_processed_page": int(last_processed_page),
        "target_table_count": len(target_tables),
        "target_tables_on_last_page": target_tables_on_last_page,
    }


def prune_cached_images(
    annotation_dict: dict,
    cached_pages: dict[int, PageImage],
    target_classes: tuple[int, ...] = (0, 1, 2, 3),
) -> dict[int, PageImage]:
    """
    Keep only rendered pages that still may be needed for TSR.
    """
    if not cached_pages:
        gc.collect()
        return {}

    target_classes_set = set(int(x) for x in target_classes)

    pages_to_keep: set[int] = set()
    for table in annotation_dict.get("tables", []):
        class_id = int(table.get("class_id", -1))
        if class_id in target_classes_set:
            pages_to_keep.add(int(table.get("page_index", -1)))

    pruned_pages: dict[int, PageImage] = {}
    for page_index, page in cached_pages.items():
        if int(page_index) in pages_to_keep:
            pruned_pages[int(page_index)] = page
        else:
            try:
                page.image_pil.close()
            except Exception:
                pass

    gc.collect()
    return pruned_pages


def run_pipeline_incremental_until_ready(
    pdf_path: str,
    final_xlsx_path: str, 
    working_dir:str,
    run_rec: bool = True,
    save_json: bool = True,
    initial_scan_pages: int = 8,
    next_scan_pages: int = 8,
    target_classes: tuple[int, ...] = (0, 1, 2, 3),
    output_dir: Optional[str] = None,
) -> dict:
    """
    Incremental orchestration:
    - process first 10 pages (or fewer if document is shorter)
    - run all early stages through table classification
    - check stop conditions
    - if not ready, prune heavy images and continue by 5 pages
    - once ready, run TSR and export Excel
    - if end of PDF is reached without readiness, still run TSR on what was found
    """
    pdf_path = str(pdf_path)
    final_xlsx_path = str(final_xlsx_path)
    working_dir = str(working_dir)
    pdf_info = pdfinfo_from_path(pdf_path)
    num_pages = int(pdf_info["Pages"])

    accumulated_annotation: Optional[dict] = None
    cached_pages: dict[int, PageImage] = {}
    next_table_id = 0

    current_start_page = 1
    current_end_page = min(initial_scan_pages, num_pages)

    readiness_info: Optional[dict] = None
    stopped_early = False

    while current_start_page <= num_pages:
 
        annotation_chunk, pages_chunk = build_annotation_with_table_ocr(
            pdf_path=pdf_path,
            run_rec=run_rec,
            first_page=current_start_page,
            last_page=current_end_page,
        )

        accumulated_annotation, cached_pages, next_table_id = merge_annotation_chunks(
            accumulated_annotation=accumulated_annotation,
            accumulated_pages=cached_pages,
            annotation_chunk=annotation_chunk,
            pages_chunk=pages_chunk,
            next_table_id=next_table_id,
        )

        accumulated_annotation, cls_results = run_table_classification_step(
            pdf_path=pdf_path,
            annotation_dict=accumulated_annotation,
            save_json=save_json,
        )

        readiness_info = check_tsr_ready_conditions(
            annotation_dict=accumulated_annotation,
            last_processed_page=current_end_page,
            target_classes=target_classes,
        )

        if readiness_info["ready_for_tsr"]:
            stopped_early = current_end_page < num_pages
            break

        cached_pages = prune_cached_images(
            annotation_dict=accumulated_annotation,
            cached_pages=cached_pages,
            target_classes=target_classes,
        )

        if current_end_page >= num_pages:
            break

        current_start_page = current_end_page + 1
        current_end_page = min(current_end_page + next_scan_pages, num_pages)

    if accumulated_annotation is None:
        accumulated_annotation = {
            "document_path": pdf_path,
            "num_pages": num_pages,
            "tables": [],
            "page_angles_deg": [0.0] * num_pages,
            "pages": [],
            "status": {},
            "logical_tables": [],
        }

    if readiness_info is None:
        readiness_info = check_tsr_ready_conditions(
            annotation_dict=accumulated_annotation,
            last_processed_page=min(current_end_page, num_pages),
            target_classes=target_classes,
        )

    pages_for_tsr = [cached_pages[k] for k in sorted(cached_pages.keys())]

    tsr_candidates = build_tsr_candidates(accumulated_annotation, pages_for_tsr)

    pipeline_results = []
    xlsx_path = None

    if len(tsr_candidates) > 0:
        pipeline_results = run_table_pipeline_on_candidates(tsr_candidates)
        xlsx_path = export_table_pipeline_results_to_excel(
            tsr_candidates=tsr_candidates,
            pipeline_results=pipeline_results,
            pdf_path=pdf_path, 
            final_xlsx_path= final_xlsx_path,
            working_dir=working_dir
        )

    if save_json:
        json_path = annotation_json_path(pdf_path)
        json_path.write_text(
            json.dumps(accumulated_annotation, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

    return {
        "pdf_path": pdf_path,
        "num_pages": num_pages,
        "processed_until_page": int(min(current_end_page, num_pages)),
        "stopped_early": bool(stopped_early),
        "ready_for_tsr": bool(readiness_info["ready_for_tsr"]),
        "readiness_info": readiness_info,
        "annotation_dict": accumulated_annotation,
        "cached_page_indices": sorted(cached_pages.keys()),
        "tsr_candidate_count": int(len(tsr_candidates)),
        "xlsx_path": str(xlsx_path) if xlsx_path is not None else None,
    }

